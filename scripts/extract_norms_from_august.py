#!/usr/bin/env python3
"""Avgust Exceldan normalarni ehtiyotkorlik bilan ajratish (faqat preview/seed)."""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = next((ROOT / "kunlik kiritish uchun malumotlar").glob("*.xlsx"))
OUT = ROOT / "scripts" / "norms_seed.json"
REPORT = ROOT / "scripts" / "_norms_report.txt"

FLEET_BY_CODE = {
    "931": "01 931 PJA",
    "668": "01 668 UKA",
    "646": "01 646 UKA",
    "449": "01 449 UKA",
    "887": "01 887 UKA",
    "269": "01 269 KMA",
    "382": "01 382 NMA",
    "870": "01 870 SEA",
    "043": "01 043 KMA",
    "849": "01 849 SNA",
    "302": "01 302 DNA",
    "949": "01 949 AKA",
    "255": "01 255 HMA",
    "205": "01 205 HMA",
    "309": "01 309 YNA",
    "592": "01 592 YNA",
    "083": "01 083 XJA",
    "282": "01 282 BMA",
    "844": "01 844 FKA",
    "331": "01 331 MLA",
    "699": "01 699 UKA",
}


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, str) and v.startswith("#"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_plate(cell):
    s = str(cell or "")
    m = re.search(r"01\s*\d{3}\s*[A-Za-zА-Яа-я]{2,3}", s)
    if not m:
        return None
    raw = m.group(0).upper()
    trans = str.maketrans("АВСЕНКМОРТХ", "ABCEHKMOPTX")
    raw = raw.translate(trans)
    parts = re.findall(r"[A-Z0-9]+", raw)
    if len(parts) >= 3:
        return f"{parts[0]} {parts[1]} {parts[2]}"
    return " ".join(parts)


def mode(vals):
    if not vals:
        return None
    rounded = [round(v, 2) for v in vals]
    return max(set(rounded), key=rounded.count)


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    out = {}
    report = [f"FILE={XLSX.name}"]

    for code, default_plate in FLEET_BY_CODE.items():
        if code not in wb.sheetnames:
            report.append(f"MISS sheet {code}")
            continue
        ws = wb[code]
        r2a = ws.cell(2, 1).value
        r2b = ws.cell(2, 5).value
        parsed = parse_plate(r2b) or parse_plate(r2a) or default_plate
        h5 = str(ws.cell(6, 5).value or "").upper()
        is_diesel = ("СОЛЯР" in h5) or ("SOLAR" in h5) or ("ДИЗЕЛ" in h5)

        gas_vals = []
        ben_vals = []
        for r in range(8, 40):
            day = ws.cell(r, 1).value
            if not isinstance(day, (int, float)):
                continue
            day = int(day)
            if day < 1 or day > 31:
                continue
            g = num(ws.cell(r, 8).value)
            b = num(ws.cell(r, 9).value)
            if g is not None and 0 < g < 40:
                gas_vals.append(g)
            if b is not None and 0 < b < 40:
                ben_vals.append(b)

        gas_n = mode(gas_vals)
        ben_n = mode(ben_vals)
        gas_start = num(ws.cell(8, 6).value)
        ben_start = num(ws.cell(8, 7).value)

        if gas_n is None and ben_n is None:
            report.append(f"NO_NORMS {code} {parsed}")
            continue

        fuel_type = "dizel_gaz" if is_diesel else "mixed"
        if gas_n is None:
            gas_n = 12.0
        if ben_n is None:
            ben_n = 4.0 if is_diesel else 7.6

        rec = {
            "sheet": code,
            "plate": parsed,
            "fuelType": fuel_type,
            "gasNorm": gas_n,
            "benzinNorm": ben_n,
            "gasStart": gas_start if gas_start is not None and gas_start >= 0 else 0,
            "benzinStart": ben_start if ben_start is not None and ben_start >= 0 else 0,
        }
        out[parsed] = rec
        report.append(
            f"OK {code} -> {parsed} gas={gas_n} ben={ben_n} type={fuel_type} "
            f"startG={rec['gasStart']} startB={rec['benzinStart']}"
        )

    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    REPORT.write_text("\n".join(report), encoding="utf-8")
    print(f"cars={len(out)}")
    print("\n".join(report))


if __name__ == "__main__":
    main()
